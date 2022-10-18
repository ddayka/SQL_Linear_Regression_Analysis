import time
import keyboard
import win32api
import win32con
import mysql.connector
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
import csv


def query(query_execute):
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="PASSWORD",
        database="psi_bw"
    )
    mycursor = db.cursor()
    mycursor.execute(query_execute)

    data = []

    for result in mycursor:
        data.append(result)

    return data


def query_to_df(query_execute):
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="PASSWORD",
        database="psi_bw"
    )
    result_df = pd.read_sql(query_execute, db)
    db.close()
    result_df.head()
    return result_df


def insert_transaction(ar, database):
    db = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="PASSWORD",
        database="psi_bw"
    )
    mycursor = db.cursor()

    mycursor.execute("INSERT INTO " + database + "(ID, Type, Status, Closes, TransactionDate, OriginalQty, "
                                                 "RemainingQty, ActualCost, InventoryCost, UnitPrice, CumulativeCost, "
                                                 "Comment) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (str(ar[0]),
                                                                                                           int(ar[1]),
                                                                                                           int(ar[2]),
                                                                                                           int(ar[3]),
                                                                                                           datetime.strptime(
                                                                                                               ar[4],
                                                                                                               '%m/%d/%Y'),
                                                                                                           int(ar[5]),
                                                                                                           int(ar[6]),
                                                                                                           float(ar[7]),
                                                                                                           float(ar[8]),
                                                                                                           float(ar[9]),
                                                                                                           float(
                                                                                                               ar[10]),
                                                                                                           str(ar[
                                                                                                                   11], )))
    db.commit()


def csv_to_array(keyword, folder_path):
    filepath = ''
    filename = ''

    for file in os.listdir(folder_path):
        if keyword in file:
            # print(file)
            filename = file
            filepath = os.path.join(folder_path, file)

    print('Path:' + str(filepath))

    # Append CSV data to an Array
    data = []

    with open(filepath) as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:  # each row is a list
            data.append(row)

    return filename, filepath, data


def df_to_excel(df, ws):
    rows = dataframe_to_rows(df)

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


def append_db(keyword, folder_path=r'DATA EXPORT FOLDERPATH'):
    #  Get File& Append CSV data to an Array

    filename, filepath, append_data = csv_to_array(keyword, folder_path)

    #  Get most recent entry in sql table
    table = "psi_bw.`" + keyword + "`"

    count = query("SELECT COUNT(*) FROM " + table)
    results = query("SELECT * FROM " + table)
    endex = int(count[0][0]) - 1

    # print("Endex: " + str(endex))
    # print("Recent Entry: " + str(results[endex]))

    #  Find Index of most recent entry in our csv array
    index = 1

    dif_bool = True

    while dif_bool:
        if results[endex][0] == str(append_data[index][0]):  # Check Part #
            # print(append_data[index])
            # print("part")
            if results[endex][4] == datetime.strptime(append_data[index][4], '%m/%d/%Y'):  # Check Date
                # print("date")
                if results[endex][11] == str(append_data[index][11]):  # Check Comment (PO # / Inv #)
                    # print("comm")
                    if results[endex][5] == int(append_data[index][5]):  # and results[endex][6] == int(append_data[index][6]):  # Check original Qty
                        # print("qty")
                        dif_bool = False
        index += 1

    # print(index)
    # print(append_data[index - 1])

    #  Add to database ascending order (Index - 1 to start, -1 until 0)
    index -= 2  # move back 1 index for duplicate pair and 1 for index 0
    while index > 0:
        insert_transaction(append_data[index], table)
        index -= 1

    # Move csv File to Archive
    os.replace(filepath, os.path.join(folder_path, str(keyword) + " Archive", filename))


def reset_db(keyword, folder_path=r'FOLDER WITH DB CSV EXPORTS'):
    table = "psi_bw.`" + keyword + "`"
    query("TRUNCATE TABLE " + table)

    filename, filepath, append_data = csv_to_array(keyword, folder_path)
    print(filename, filepath)

    db = mysql.connector.connect(
        host="localhost",
        user="root",
        passwd="PASSWORD",
        database="psi_bw"
    )
    mycursor = db.cursor()
    
    # Update sql Database from CSV Files downloaded from accounting software
    for entry in append_data[1:]:
        mycursor.execute(
            "INSERT INTO " + table + "(ID, OnHandQty, OnOrderQty, CommittedQty, MTDReceipts, MTDIssues, MTDAdjust, "
                                     "MTDSales, MTDCostOfGoods, YTDReceipts, YTDIssues, YTDAdjust, YTDSales, "
                                     "YTDCostOfGoods, PriorYTDReceipts, PriorYTDIssues, PriorYTDAdjust, "
                                     "PriorYTDSales, PriorYTDCostOfGoods ) VALUES "
                                     "(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (
                str(entry[0]),  # ID
                int(round(float(entry[1]), 0)),  # OnHandQTY
                int(entry[2]),  # OnOrderQTY
                int(entry[3]),  # CommittedQTY
                int(entry[4]),  # MTD Receipts
                int(entry[5]),  # MTD Issues
                int(entry[6]),  # MTD Adjust
                float(entry[7]),  # MTD Sales
                float(entry[8]),  # MTD COGS
                int(entry[9]),  # YTD Reciepts
                int(entry[10]),  # YTD Issues
                int(entry[11]),  # YTD Adjusts
                float(entry[12]),  # YTD Sales
                float(entry[13]),  # MTD COGS
                int(entry[14]),  # PYTD Reciepts
                int(entry[15]),  # PYTD Issues
                int(entry[16]),  # PYTD Adjusts
                float(entry[17]),  # PYTD Sales
                float(entry[18])  # PMTD COGS
            )
        )

    db.commit()

    os.replace(filepath, os.path.join(folder_path, str(keyword) + " Archive", filename))


def inventory_query(where_vendor=""):
    def if_where_like(column_id, param):
        if param != "":
            return " WHERE " + column_id + " LIKE " + "'" + param + "%'"
        else:
            return ""

    query_str = """
        # SET @tValue = 0.978; #80% T Value @ N=3
        # SET @N = 3; #Obeservations - 2
        
        SELECT ID, `DESCRIPTION`, QUANTITY, 
        AVG_ANNUAL_SALE, SALE_21, SALE_20,
        AVG_SERV_QTY, MaxService, MinService,
        VENDOR_ID, AVG_ANNUAL_NUM_RECEIPT, NUM_RCPT_21, NUM_RCPT_20
        FROM (
            SELECT i.ID, i. Description1 as `DESCRIPTION`, (p.OnHandQty + p.OnOrderQty - p.CommittedQty) AS QUANTITY, 
            # Annual Inventory Issues
            round((p.OnHandQty + p.OnOrderQty - p.CommittedQty)/UP_DAILY, 0) AS UP_DAY_REMAIN, round((p.OnHandQty + p.OnOrderQty - p.CommittedQty)/LOW_DAILY, 0) AS LOW_DAY_REMAIN,
            AVG_ANNUAL_SALE, UP_ANNUAL, LOW_ANNUAL, UpperBoundAVG, LowBoundAVG,
            SALE_17, SALE_18, SALE_19, SALE_20, SALE_21,
            Slope, Intercept, R2, ROUND(StdError, 2) as StdEr,
            # Average Issues per Service
            AvgService AS AVG_SERV_QTY, VarService, MaxService, MinService,
            # Annual Inventory Reciepts
            i.APVendorID as VENDOR_ID, i.RemitName as VENDOR_NAME,
            AVG_ANNUAL_RECEIPT, RCPT_17, RCPT_18, RCPT_19, RCPT_20, RCPT_21, 
            AVG_ANNUAL_NUM_RECEIPT, NUM_RCPT_17, NUM_RCPT_18, NUM_RCPT_19, NUM_RCPT_20, NUM_RCPT_21 
            FROM psi_bw.`psi ic parts_on_hand` as p
            RIGHT JOIN (
                SELECT *,
                UpperBound/365 AS UP_DAILY, UpperBound as UP_ANNUAL,
                LowBound/365 AS LOW_DAILY, LowBound as LOW_ANNUAL
                FROM (
                    SELECT *,
                    ROUND(1-(SSr/SSt), 4) AS R2,
                    ROUND((6 * (Slope + (0.978 * StdError)) + Intercept), 2) As UpperBound,
                    ROUND((6 * (Slope - (0.978 * StdError)) + Intercept), 2) AS LowBound,
                    ROUND(AVG_ANNUAL_SALE + (1.638 * StdError), 2) As UpperBoundAVG,
                    ROUND(AVG_ANNUAL_SALE - (1.638 * StdError), 2) AS LowBoundAVG
                    FROM (    
                        SELECT *,
                        S / sqrt(SSxx) AS StdError,
                        power(SALE_17 - AVG_ANNUAL_SALE, 2) + power(SALE_18 - AVG_ANNUAL_SALE, 2) + power(SALE_19 - AVG_ANNUAL_SALE, 2) + power(SALE_20 - AVG_ANNUAL_SALE, 2) + power(SALE_21 - AVG_ANNUAL_SALE, 2) AS SST,
                        power(SALE_17 - ((Slope*1)+Intercept), 2) + power(SALE_18 - ((Slope*2)+Intercept), 2) + power(SALE_19 - ((Slope*3)+Intercept), 2) + power(SALE_20 - ((Slope*4)+Intercept), 2) + power(SALE_21 - ((Slope*5)+Intercept), 2) AS SSR 
                        FROM (
                            SELECT *,
                            (Sy - Slope * Sx)  / 5 AS Intercept,
                            sqrt((SSyy - ((SSxy/SSxx) * SSxy)) / (3)) AS S,
                            SSyy - ((SSxy/SSxx)*SSxy)  as SSE
                            FROM (
                                SELECT *,
                                ((5 * Sxy) - (Sx * Sy)) / (5 * Sx2 - (power(Sx, 2))) as Slope,
                                Sx2 - (power(Sx, 2)/5) as SSxx,
                                Sy2 - (power(Sy, 2)/5) as SSyy,
                                Sxy - ((Sx*Sy)/5) as SSxy
                                FROM (
                                    SELECT *,
                                    ROUND((SALE_17 + SALE_18 + SALE_19 + SALE_20 + SALE_21)/5,2) as AVG_ANNUAL_SALE,
                                    15 as Sx,
                                    (SALE_17 + SALE_18 + SALE_19 + SALE_20 + SALE_21) AS Sy,
                                    (SALE_17 * 1 + SALE_18 * 2 + SALE_19 * 3 + SALE_20 * 4 + SALE_21 * 5) AS Sxy,
                                    55 as Sx2,
                                    (SALE_17 * SALE_17 + SALE_18 * SALE_18 + SALE_19 * SALE_19 + SALE_20 * SALE_20 + SALE_21 * SALE_21) AS Sy2
                                    FROM (
                                        SELECT p.*, IFNULL(d17.sale, 0) as SALE_17, IFNULL(d18.sale, 0) as SALE_18, IFNULL(d19.sale, 0) as SALE_19, IFNULL(d20.sale, 0) as SALE_20, IFNULL(d21.sale, 0) as SALE_21
                                        FROM (
                                            (
                                                SELECT *
                                                FROM psi_bw.`psi ic parts`) as p
                                            LEFT JOIN (
                                                Select ID, sum(OriginalQty) as sale
                                                FROM psi_bw.`psi ic issues`
                                                WHERE TransactionDate BETWEEN '2017-01-01' AND '2017-12-31'
                                                GROUP BY ID) AS d17
                                            ON p.ID = d17.ID
                                            LEFT JOIN (
                                                Select ID, sum(OriginalQty) as sale
                                                FROM psi_bw.`psi ic issues`
                                                WHERE TransactionDate BETWEEN '2018-01-01' AND '2018-12-31'
                                                GROUP BY ID ) AS d18
                                            ON p.ID = d18.ID
                                            LEFT JOIN ( 
                                                Select ID, sum(OriginalQty) as sale
                                                FROM psi_bw.`psi ic issues`
                                                WHERE TransactionDate BETWEEN '2019-01-01' AND '2019-12-31'
                                                GROUP BY ID) AS d19
                                            ON p.ID = d19.ID
                                            LEFT JOIN (
                                                Select ID, sum(OriginalQty) as sale
                                                FROM psi_bw.`psi ic issues`
                                                WHERE TransactionDate BETWEEN '2020-01-01' AND '2020-12-31'
                                                GROUP BY ID) AS d20
                                            ON p.ID = d20.ID
                                            LEFT JOIN (
                                                Select ID, sum(OriginalQty) as sale
                                                FROM psi_bw.`psi ic issues`
                                                WHERE TransactionDate BETWEEN '2021-01-01' AND '2021-12-31'
                                                GROUP BY ID) AS d21
                                            ON p.ID = d21.ID
                                        ) 
                                    ) AS q
                                    WHERE (SALE_17 != 0 OR SALE_18 != 0 OR SALE_19 != 0 OR SALE_20 != 0 OR SALE_21 != 0) 
                                    AND (ID NOT LIKE 'TST%' AND ID NOT LIKE 'LBL%' AND ID NOT LIKE 'LRS%' AND ID NOT LIKE 'BRP%' AND ID NOT LIKE 'CYL%' AND ID NOT LIKE 'ISS%')
                                ) AS sum
                            ) AS ss
                        ) AS StdEr
                    ) AS Bound
                ) AS R2
            ) AS i
            ON p.ID = i.ID
            LEFT JOIN (
                SELECT *,
                ROUND((RCPT_17 + RCPT_18 + RCPT_19 + RCPT_20 + RCPT_21)/5, 2) as AVG_ANNUAL_RECEIPT,
                ROUND((NUM_RCPT_17 + NUM_RCPT_18 + NUM_RCPT_19 + NUM_RCPT_20 + NUM_RCPT_21)/5, 2) as AVG_ANNUAL_NUM_RECEIPT
                FROM (
                    SELECT p.ID, p.APVendorID, IFNULL(d17.RCPT, 0) as RCPT_17, IFNULL(d18.RCPT, 0) as RCPT_18, IFNULL(d19.RCPT, 0) as RCPT_19, IFNULL(d20.RCPT, 0) as RCPT_20, IFNULL(d21.RCPT, 0) as RCPT_21,
                    IFNULL(d17.Annual_RCPT, 0) as NUM_RCPT_17, IFNULL(d18.Annual_RCPT, 0) as NUM_RCPT_18, IFNULL(d19.Annual_RCPT, 0) as NUM_RCPT_19, IFNULL(d20.Annual_RCPT, 0) as NUM_RCPT_20, IFNULL(d21.Annual_RCPT, 0) as NUM_RCPT_21
                    FROM ((
                            SELECT *
                            FROM psi_bw.`psi ic parts`) as p
                        LEFT JOIN (
                            Select ID, sum(OriginalQty) as RCPT, Count(*) as Annual_RCPT
                            FROM psi_bw.`psi ic receipts`
                            WHERE TransactionDate BETWEEN '2017-01-01' AND '2017-12-31'
                            GROUP BY ID) AS d17
                        ON p.ID = d17.ID
                        LEFT JOIN (
                            Select ID, sum(OriginalQty) as RCPT, Count(*) as Annual_RCPT
                            FROM psi_bw.`psi ic receipts`
                            WHERE TransactionDate BETWEEN '2018-01-01' AND '2018-12-31'
                            GROUP BY ID ) AS d18
                        ON p.ID = d18.ID
                        LEFT JOIN ( 
                            Select ID, sum(OriginalQty) as RCPT, Count(*) as Annual_RCPT
                            FROM psi_bw.`psi ic receipts`
                            WHERE TransactionDate BETWEEN '2019-01-01' AND '2019-12-31'
                            GROUP BY ID) AS d19
                        ON p.ID = d19.ID
                        LEFT JOIN (
                            Select ID, sum(OriginalQty) as RCPT, Count(*) as Annual_RCPT
                            FROM psi_bw.`psi ic receipts`
                            WHERE TransactionDate BETWEEN '2020-01-01' AND '2020-12-31'
                            GROUP BY ID) AS d20
                        ON p.ID = d20.ID
                        LEFT JOIN (
                            Select ID, sum(OriginalQty) as RCPT, Count(*) as Annual_RCPT
                            FROM psi_bw.`psi ic receipts`
                            WHERE TransactionDate BETWEEN '2021-01-01' AND '2021-12-31'
                            GROUP BY ID) AS d21
                        ON p.ID = d21.ID) 
                ) AS q
                WHERE (RCPT_17 != 0 OR RCPT_18 != 0 OR RCPT_19 != 0 OR RCPT_20 != 0 OR RCPT_21 != 0) 
                AND (ID NOT LIKE 'TST%' AND ID NOT LIKE 'LBL%' AND ID NOT LIKE 'LRS%' AND ID NOT LIKE 'BRP%' AND ID NOT LIKE 'CYL%' AND ID NOT LIKE 'ISS%')
            ) AS r
            ON i.ID = r.ID
            LEFT JOIN (
                SELECT ID, count(*) as TotalService, avg(OriginalQty) as AvgService, ROUND(VARIANCE(OriginalQty), 2) as VarService, max(OriginalQty) MaxService, min(OriginalQty) MinService
                FROM (
                    SELECT a.*, b.ID as CustID, b.name as CustName
                    FROM (
                        SELECT*, CONVERT(SUBSTRING(Comment, 11, 15), UNSIGNED INTEGER) AS Invoice
                        FROM psi_bw.`psi ic issues`
                        ) AS a
                    LEFT JOIN (
                        SELECT ID, Name, InvoiceNo
                        FROM psi_bw.`psi oe invoices`
                        ) AS b
                    ON a.Invoice = b.InvoiceNo
                    # WHERE b.ID NOT LIKE 'WSF'
                ) as fix
                GROUP BY ID
            ) AS o
            ON i.ID = o.ID
        ) AS Final
    """ + if_where_like("VENDOR_ID", where_vendor)

    return query_str


# MAIN

dl = True
update_db = True

# Export to CSV
if dl:
    download_file()

# Update SQL Databases
if update_db:
    append_db('PSI IC ISSUES')
    append_db('PSI IC RECEIPTS')
    append_db('PSI IC ADJUSTMENTS')
    reset_db('PSI IC PARTS_ON_HAND')

# RUN INVENTORY QUERY & EXPORT RESULTS TO EXCEL

date_today = datetime.today().strftime('%Y-%m-%d')

filepath = "EXPORT FILEPATH/FILENAME_" + date_today + ".xlsx"

wb = openpyxl.Workbook()
ws1 = wb.create_sheet("Sheet_1")
ws1.title = "Inv_All"
wb.remove(wb["Sheet"])

wb.save(filepath)


def query_to_excel(query_execute, workbook, sheet, new_sheet=False):
    query_result = query_to_df(query_execute)
    if new_sheet:
        nws = wb.create_sheet(sheet)
        nws.title = sheet
    ws = workbook[sheet]
    df_to_excel(query_result, ws)
    wb.save(filepath)


query_to_excel(inventory_query(), wb, "Inv_All")

query_to_excel(inventory_query(where_vendor='RFD'), wb, "Inv_Survitec", new_sheet=True)

query_to_excel(inventory_query(where_vendor='LSA'), wb, "Inv_LSA", new_sheet=True)

query_to_excel(inventory_query(where_vendor='ACR'), wb, "Inv_ACR", new_sheet=True)
